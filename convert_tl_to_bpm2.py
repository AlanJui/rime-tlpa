#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re

import openpyxl

excel_file = '漢字閩南語標音【台羅拼音】.xlsx'
wb = openpyxl.load_workbook(excel_file)

# === 1. 解析規則表，建立初／韻母對應字典 ===
rules_ws = wb['台羅轉換注音二式規則']
# 找到 header
header_row = None
for row in rules_ws.iter_rows(min_row=1, max_row=20):
    if any(cell.value == '台羅拼音' for cell in row):
        header_row = row[0].row
        break
if header_row is None:
    raise RuntimeError('規則表找不到標頭「台羅拼音」')

# 拿到各欄的 index
cols = {cell.value: cell.column for cell in rules_ws[header_row]}
tl_col_idx     = cols['台羅拼音']
bpm2_col_idx   = cols['注音二式']
kind_col_idx   = cols['音標種類']  # 新增欄：聲母 or 韻母

initial_map = {}
final_map   = {}
for row in rules_ws.iter_rows(min_row=header_row+1, max_row=rules_ws.max_row):
    key  = row[tl_col_idx-1].value
    val  = row[bpm2_col_idx-1].value
    kind = row[kind_col_idx-1].value
    if key and val and kind:
        key = str(key).strip()
        val = str(val).strip()
        if kind.strip() == '聲母':
            initial_map[key] = val
        elif kind.strip() == '韻母':
            final_map[key] = val

# 為了 longest‐match，先把聲母鍵由長到短排序
initial_keys = sorted(initial_map.keys(), key=len, reverse=True)

# === 2. 處理 tl_ji_khoo_phing 工作表 ===
ws = wb['tl_ji_khoo_phing']
# 找 header row 與所需欄位
hdr = next(ws.iter_rows(min_row=1, max_row=1))
header_cols = {cell.value: cell.column for cell in hdr}

if '台羅拼音' not in header_cols or '注音二式' not in header_cols:
    raise RuntimeError('tl_ji_khoo_phing 表缺「台羅拼音」或「注音二式」')

src_col  = header_cols['台羅拼音']
dst_col  = header_cols['注音二式']

# 如果沒有「訂正說明」欄，就在最右邊新增一欄
if '訂正說明' not in header_cols:
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value='訂正說明')
    note_col = new_col
else:
    note_col = header_cols['訂正說明']

def convert_one(syl):
    """將一個台羅音節轉為注音二式。"""
    if syl is None:
        return ''
    s = str(syl).strip()
    # 分離尾端數字
    m = re.match(r'^(.+?)([0-9])$', s)
    body, tone = (m.group(1), m.group(2)) if m else (s, '')
    # 找最長匹配的聲母
    init = ''
    for k in initial_keys:
        if body.startswith(k):
            init = k
            break
    final = body[len(init):]

    # Fallback：找不到就先用原字串
    init_p  = initial_map.get(init, init)
    final_p = final_map.get(final, final)

    # debug 輸出：細分哪個沒對上
    if not init_p and not final_p:
        print(f'⚠️ 無法轉換：{syl}')
        return ''
    else:
        print(f'syl: {syl}, init: {init} -> {init_p}, final: {final} -> {final_p}, tone: {tone}')

    missing = []
    if init not in initial_map:
        missing.append('聲母未轉換')
    if final not in final_map:
        missing.append('韻母未轉換')
    if missing:
        note = '、'.join(missing)
        print(f'⚠️ {note}：{syl} → init:{init}->{init_p}, final:{final}->{final_p}, tone:{tone}')

    return init_p + final_p + tone


# # 逐列從第 2 到第 101 列
# for r in range(2, 102):
#     cell_src  = ws.cell(row=r, column=src_col).value
#     result    = convert_one(cell_src)
#     ws.cell(row=r, column=dst_col, value=result)
#     if not result:
#         # 沒對到就註記
#         ws.cell(row=r, column=note_col,
#                 value=f'缺規則：{cell_src}')
# 轉換全部的【漢字標音】：逐列從第 2 列到最後一列
for r in range(2, ws.max_row + 1):
    cell_src  = ws.cell(row=r, column=src_col).value
    result    = convert_one(cell_src)
    ws.cell(row=r, column=dst_col, value=result)
    if not result:
        # 沒對到就註記
        ws.cell(row=r, column=note_col,
                value=f'缺規則：{cell_src}')


# === 3. 存檔 ===
wb.save(excel_file)
print('✅ 完成：前 100 列轉換並標註「訂正說明」。')
