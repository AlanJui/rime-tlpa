#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
a872_BPM2字典切台羅韻調聲.py

將【台語注音二式字庫】.xlsx 的【台羅拼音】工作表中，
code（B欄）拆成【韻 + 調 + 聲】，分別寫入：
  I欄＝韻母
  J欄＝聲調
  K欄＝聲母（零聲母填 q）

用法：
  py -3 tools/a872_BPM2字典切台羅韻調聲.py
  py -3 tools/a872_BPM2字典切台羅韻調聲.py path/to/file.xlsx
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

# 台羅聲母（最長優先）
# 含齒音連 i：tshi / tsi / ji / si
TL_INITIALS = (
    "tshi",
    "tsi",
    "tsh",
    "ji",
    "si",
    "ts",
    "ph",
    "th",
    "kh",
    "ng",
    "m",
    "b",
    "p",
    "n",
    "l",
    "t",
    "g",
    "k",
    "j",
    "s",
    "h",
)

CODE_RE = re.compile(r"^([a-z]+)(\d+)$")
# 韻化聲母（整音節僅 m / ng + 調）
SYLLABIC_NASALS = frozenset({"m", "ng"})

DEFAULT_XLSX = Path(__file__).resolve().parent.parent / "src" / "【台語注音二式字庫】.xlsx"
SHEET_NAME = "台羅拼音"

# Excel 欄位：I=9 韻, J=10 調, K=11 聲
COL_FINAL = 9
COL_TONE = 10
COL_INITIAL = 11


def split_tl_code(code: str) -> tuple[str, str, str] | None:
    """回傳 (聲, 韻, 調)；無法解析則回傳 None。"""
    m = CODE_RE.match(code.strip().lower())
    if not m:
        return None
    body, tone = m.group(1), m.group(2)

    if body in SYLLABIC_NASALS:
        return "q", body, tone

    for ini in TL_INITIALS:
        if body.startswith(ini) and len(body) > len(ini):
            return ini, body[len(ini) :], tone

    # 零聲母
    return "q", body, tone


def process_workbook(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"找不到工作表：{SHEET_NAME}")

    ws = wb[SHEET_NAME]
    ws.cell(1, COL_FINAL).value = "韻"
    ws.cell(1, COL_TONE).value = "調"
    ws.cell(1, COL_INITIAL).value = "聲"

    ok = 0
    skipped = 0
    zero_initial = 0

    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row, 2).value  # B = code
        if raw is None or str(raw).strip() == "":
            skipped += 1
            continue

        code = str(raw).strip()
        parts = split_tl_code(code)
        if parts is None:
            print(f"  [略過] 第 {row} 列無法解析 code={code!r}")
            skipped += 1
            continue

        siann, un, tiau = parts
        ws.cell(row, COL_FINAL).value = un
        ws.cell(row, COL_TONE).value = tiau
        ws.cell(row, COL_INITIAL).value = siann
        ok += 1
        if siann == "q":
            zero_initial += 1

    wb.save(xlsx_path)
    print(f"已寫入：{xlsx_path}")
    print(f"工作表【{SHEET_NAME}】：成功 {ok} 列，略過 {skipped} 列，零聲母(q) {zero_initial} 列")


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        raise SystemExit(f"找不到檔案：{xlsx}")
    process_workbook(xlsx)


if __name__ == "__main__":
    main()
