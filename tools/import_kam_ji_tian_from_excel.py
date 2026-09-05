#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
import_kam_ji_tian_from_excel.py

功能：
  依《420_甘字典漢字庫指引.md》規格，自 Excel
  【甘字典】ChhoeTaigi_KamJitian.xlsx 之【RIME】工作表，
  將資料匯入 SQLite【甘字典漢字庫】（漢字庫資料表）。

欄位對映：
  A text   → 漢字
  B code   → 台羅音標
  C weight → 常用度
  D stem   → 摘要說明
  （時間戳記由資料庫 DEFAULT / 觸發器處理；若 E 欄 create 有值則一併寫入）

使用方式：
  python tools/import_kam_ji_tian_from_excel.py
  python tools/import_kam_ji_tian_from_excel.py --xlsx path/to.xlsx
  python tools/import_kam_ji_tian_from_excel.py --db src/Kam_Ji_Tian.db --replace
  python tools/import_kam_ji_tian_from_excel.py --selftest
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))

from kam_ji_tian_db import DEFAULT_DB, init_db, connect  # noqa: E402

DEFAULT_XLSX = ROOT / "src" / "【甘字典】ChhoeTaigi_KamJitian.xlsx"
SHEET_NAME = "RIME"
START_ROW = 2  # 第 1 列為標題


def _cell_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _cell_weight(value) -> float:
    if value is None or str(value).strip() == "":
        return 0.1
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.1


def _normalize_create(value) -> str | None:
    """將 Excel create 欄轉成 'YYYY-MM-DD HH:MM:SS'；空值回傳 None。"""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    s = str(value).strip().replace("/", "-")
    # 允許未補零日期：2026-7-5 21:19
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    # 手動解析未補零：YYYY-M-D[ HH:MM[:SS]]
    try:
        date_part, _, time_part = s.partition(" ")
        y, m, d = date_part.split("-")
        hh = mm = ss = 0
        if time_part:
            parts = time_part.split(":")
            hh = int(parts[0])
            mm = int(parts[1]) if len(parts) > 1 else 0
            ss = int(parts[2]) if len(parts) > 2 else 0
        dt = datetime(int(y), int(m), int(d), hh, mm, ss)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return s


def parse_rime_row(row) -> tuple[str, str, float, str, str | None] | None:
    """
    解析 RIME 工作表一列。
    回傳 (漢字, 台羅音標, 常用度, 摘要說明, 時間戳記|None)；無效列回傳 None。
    """
    if row is None:
        return None
    cells = list(row) if not isinstance(row, (str, bytes)) else [row]
    while len(cells) < 5:
        cells.append(None)

    text = _cell_str(cells[0])
    code = _cell_str(cells[1]).lower()
    if not text or not code:
        return None
    if text in ("-", "?", "？", "text"):
        return None

    weight = _cell_weight(cells[2])
    stem = _cell_str(cells[3]) or "NA"
    create = _normalize_create(cells[4])
    return text, code, weight, stem, create


def read_rime_sheet(xlsx_path: Path) -> list[tuple]:
    """以 openpyxl 讀取【RIME】工作表（不需 Excel 常駐）。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("✗ 未安裝 openpyxl，請先執行：pip install openpyxl")

    if not xlsx_path.is_file():
        sys.exit(f"✗ 找不到檔案：{xlsx_path}")

    print(f"開啟活頁簿：{xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"✗ 找不到工作表：{SHEET_NAME}（現有：{wb.sheetnames}）")

    ws = wb[SHEET_NAME]
    records: list[tuple] = []
    for idx, row in enumerate(ws.iter_rows(min_row=START_ROW, values_only=True), start=START_ROW):
        parsed = parse_rime_row(row)
        if parsed is None:
            continue
        text, code, weight, stem, create = parsed
        records.append((idx, text, code, weight, stem, create))
        print(f"\r目前處理 列號={idx} 漢字={text} 台羅音標={code}", end="", flush=True)
    print()
    wb.close()
    return records


def import_records(
    records: list[tuple],
    db_path: Path,
    *,
    replace: bool = False,
) -> tuple[int, int, int]:
    """
    寫入【漢字庫】。
    回傳 (寫入筆數, 更新筆數, 略過筆數)。
    以 UNIQUE(漢字, 台羅音標) 處理重複：replace=True 時 UPSERT，否則略過。
    """
    init_db(db_path)
    inserted = updated = skipped = 0
    conn = connect(db_path)
    try:
        if replace:
            # 清空後全量匯入（符合「自工作表重建字庫」情境）
            conn.execute("DELETE FROM 漢字庫")
            conn.commit()

        sql_upsert = """
            INSERT INTO 漢字庫 (漢字, 台羅音標, 常用度, 摘要說明, 時間戳記)
            VALUES (?, ?, ?, ?, COALESCE(?, DATETIME('now', 'localtime')))
            ON CONFLICT(漢字, 台羅音標) DO UPDATE SET
                常用度 = excluded.常用度,
                摘要說明 = excluded.摘要說明,
                時間戳記 = COALESCE(excluded.時間戳記, 時間戳記)
        """

        for row_no, text, code, weight, stem, create in records:
            params = (text, code, weight, stem, create)
            try:
                exists = conn.execute(
                    "SELECT 1 FROM 漢字庫 WHERE 漢字=? AND 台羅音標=?",
                    (text, code),
                ).fetchone()
                # 全量／增量皆用 UPSERT：同鍵重複時保留最後一筆
                conn.execute(sql_upsert, params)
                if exists:
                    updated += 1
                else:
                    inserted += 1
            except sqlite3.IntegrityError:
                skipped += 1
            print(
                f"\r寫入資料庫 列號={row_no} 漢字={text} 台羅音標={code}",
                end="",
                flush=True,
            )
        print()
        conn.commit()
    finally:
        conn.close()

    return inserted, updated, skipped


def selftest() -> int:
    ok = True

    def check(desc, got, exp):
        nonlocal ok
        mark = "OK  " if got == exp else "FAIL"
        if got != exp:
            ok = False
        print(f"{mark} {desc}: {got!r} (expect {exp!r})")

    check(
        "正常列",
        parse_rime_row(["了", "liau2", 0.80, "【13769】定tio̍h.", "2026/7/5 21:19"]),
        ("了", "liau2", 0.80, "【13769】定tio̍h.", "2026-07-05 21:19:00"),
    )
    check("缺漢字 → None", parse_rime_row([None, "liau2", 0.8, "x", None]), None)
    check("缺音標 → None", parse_rime_row(["了", "", 0.8, "x", None]), None)
    check("標題列 → None", parse_rime_row(["text", "code", "weight", "stem", "create"]), None)
    check(
        "weight 空 → 0.1",
        parse_rime_row(["一", "it4", None, "NA", None])[2],
        0.1,
    )
    check(
        "stem 空 → NA",
        parse_rime_row(["一", "it4", 0.8, None, None])[3],
        "NA",
    )

    # 暫存 DB 驗證 DDL + UPSERT（Windows 需 ignore_cleanup_errors）
    import tempfile

    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
        db = Path(tmp) / "test.db"
        init_db(db)
        recs = [
            (2, "了", "liau2", 0.80, "【13769】a", "2026-07-05 21:19:00"),
            (3, "了", "liau2", 0.60, "【13769】b", "2026-07-05 21:20:00"),
            (4, "一", "it4", 0.80, "【8221】c", None),
        ]
        ins, upd, sk = import_records(recs, db, replace=False)
        check("UPSERT insert+update", (ins, upd, sk), (2, 1, 0))
        conn = connect(db)
        try:
            n = conn.execute("SELECT COUNT(*) FROM 漢字庫").fetchone()[0]
            stem = conn.execute(
                "SELECT 摘要說明 FROM 漢字庫 WHERE 漢字=? AND 台羅音標=?",
                ("了", "liau2"),
            ).fetchone()[0]
        finally:
            conn.close()
        check("UPSERT 後筆數", n, 2)
        check("UPSERT 後摘要", stem, "【13769】b")

    print("ALL PASS" if ok else "SOME FAILED")
    return 0 if ok else 1


def main() -> None:
    ap = argparse.ArgumentParser(description="【RIME】工作表 → 甘字典漢字庫（SQLite）")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="輸入 .xlsx")
    ap.add_argument("--db", default=str(DEFAULT_DB), help="SQLite 資料庫路徑")
    ap.add_argument(
        "--replace",
        action="store_true",
        help="匯入前清空【漢字庫】後全量寫入（預設為 UPSERT）",
    )
    ap.add_argument("--selftest", action="store_true", help="僅執行自我測試")
    args = ap.parse_args()

    if args.selftest:
        sys.exit(selftest())

    xlsx = Path(args.xlsx)
    db_path = Path(args.db)
    records = read_rime_sheet(xlsx)
    if not records:
        sys.exit("✗ 【RIME】工作表無有效資料可匯入。")

    print(f"準備寫入資料庫：{db_path}（共 {len(records)} 筆）")
    inserted, updated, skipped = import_records(
        records, db_path, replace=args.replace
    )
    mode = "全量重建" if args.replace else "UPSERT"
    print(
        f"✓ 完成（{mode}）：寫入 {inserted} 筆，更新 {updated} 筆，略過 {skipped} 筆 → {db_path}"
    )


if __name__ == "__main__":
    main()
