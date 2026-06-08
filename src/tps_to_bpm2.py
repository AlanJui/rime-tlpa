#!/usr/bin/env python3
"""
tps_to_bpm2.py v0.0.3 — 方音符號 (TPS) → 台語注音二式 (BPM2) 批量轉換

讀取 CIN方音符號.xlsx 中的 【聲】聲母對照表、【韻】韻母對照表，
將【漢字庫】工作表 E 欄（方音符號）批量轉換後寫入 F 欄（台語注音二式）。

【註】：執行前，需將 Excel 活頁簿檔案關閉，否則 openpyxl 無法寫入。

用法：
    python tps_to_bpm2.py [excel_path] [output_path]

    excel_path  : 來源 .xlsx（預設同目錄的 CIN方音符號.xlsx）
    output_path : 輸出 .xlsx（預設覆寫來源檔）

(.venv) PS C:\Users\AlanJui\work\rime-tlpa\src> py .\tps_to_bpm2.py
讀取：C:\Users\AlanJui\work\rime-tlpa\src\CIN方音符號.xlsx
  聲母對映：21 筆
  韻母對映：162 筆
  按鍵對映：49 筆

開始轉換（共約 21,026 列）…
  處理中… row 21,029  (100.0%)

轉換完成：
  成功 21,023 筆 / 失敗 2 筆（涵蓋率 100.0%）
  空白略過 1 筆

失敗明細（共 2 筆）：
  row  2,078  key='qjj2'  TPS='ㄆㆫㄉ'
  row  2,079  key='qjj2'  TPS='ㄆㆫㄉ'

輸出：C:\Users\AlanJui\work\rime-tlpa\src\CIN方音符號.xlsx
(.venv) PS C:\Users\AlanJui\work\rime-tlpa\src>
"""

import sys

import pandas as pd
from openpyxl import load_workbook

# ── 1. 音調末字元對映 ─────────────────────────────────────────────────────────
# last codepoint → BPM2 tone digit
TONE_SUFFIX: dict[str, str] = {
    '˫': '7',   # ˫  U+02EB
    '˪': '3',   # ˪  U+02EA
    'ˋ': '2',   # ˋ  U+02CB
    'ˊ': '5',   # ˊ  U+02CA
    '˙': '8',   # ˙  U+02D9
    '̀': '2',   # combining grave (alternate)
    '́': '5',   # combining acute (alternate)
}

# 入聲尾輔音（用於判斷調 4；˙ 後接此類字元 → 調 8 已先處理）
JIP_SIANN_UN_BUE: set[str] = {
    'ㄏ',   # ㄏ h-final
    'ㄍ',   # ㄍ k-final
    'ㄉ',   # ㄉ t-final
    'ㄅ',   # ㄅ p-final
}

# ── 2. 鍵盤雙字元產生的 TPS 重複字元 → 單一 TPS 符號 ────────────────────────
# 來源：按鍵對映工作表（'ss'→ㄫ 等），Excel 公式逐字元展開造成重複
DOUBLE_TPS_FIX: dict[str, str] = {
    'ㄋㄋ': 'ㄫ',   # ㄋㄋ → ㄫ  (ss → ng 聲母, U+312B)
    'ㄚㄚ': 'ㆩ',   # ㄚㄚ → ㆩ  (88 → ann)
    'ㄞㄞ': 'ㆮ',   # ㄞㄞ → ㆮ  (99 → ainn)
    'ㆤㆤ': 'ㆥ',   # ㆤㆤ → ㆥ  (,, → enn)
    'ㄛㄛ': 'ㆧ',   # ㄛㄛ → ㆧ  (ii → oonn)
    'ㄨㄨ': 'ㆫ',   # ㄨㄨ → ㆫ  (jj → nn, U+31AB)
    'ㄠㄠ': 'ㆯ',   # ㄠㄠ → ㆯ  (ll → iaunn)
    'ㄧㄧ': 'ㆪ',   # ㄧㄧ → ㆪ  (uu → inn)
}


def _normalize_tps(s: str) -> str:
    """修正 Excel 公式逐字元展開造成的重複 TPS 符號。"""
    for dbl, single in DOUBLE_TPS_FIX.items():
        s = s.replace(dbl, single)
    return s


# ── 3. 建立轉換表 ─────────────────────────────────────────────────────────────
def build_tables(path: str):
    """
    回傳：
        siann_map : {TPS聲母字元 → BPM2聲母字串}
        siann_set : TPS 聲母字元集合
        fin_map   : {TPS韻母字串 → BPM2韻母字串}
    """
    # ── 聲母 ──
    ds = pd.read_excel(path, sheet_name='【聲】聲母對照表', header=0)
    siann_map: dict[str, str] = {}
    siann_set: set[str] = set()
    for _, row in ds.dropna(subset=['方音符號']).iterrows():
        tps  = str(row['方音符號']).strip()
        bpm2 = str(row['台語注音二式']).strip() if pd.notna(row['台語注音二式']) else ''
        if tps and tps != 'nan':
            siann_map[tps] = bpm2
            siann_set.add(tps)

    # ── 韻母 ──
    dk_raw = pd.read_excel(path, sheet_name='【韻】韻母對照表', header=None)
    # Row 0–1: title/blank; Row 2: header; Row 3+: data
    fin_map: dict[str, str] = {}
    for _, row in dk_raw.iloc[3:].iterrows():
        tps  = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else ''
        bpm2 = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
        if tps and bpm2 and tps != 'nan' and bpm2 != 'nan':
            fin_map[tps] = bpm2

    # 「入聲尾」別名：韻母表用特殊字元（ㆷㆻㆵㆴ），實際資料用一般字元（ㄏㄍㄉㄅ）
    SPECIAL_TO_REG = {
        'ㆷ': 'ㄏ',   # ㆷ → ㄏ
        'ㆻ': 'ㄍ',   # ㆻ → ㄍ
        'ㆵ': 'ㄉ',   # ㆵ → ㄉ
        'ㆴ': 'ㄅ',   # ㆴ → ㄅ
    }
    for tps, bpm2 in list(fin_map.items()):
        variant = ''.join(SPECIAL_TO_REG.get(c, c) for c in tps)
        if variant != tps:
            fin_map[variant] = bpm2

    # 額外別名
    # ㄛ (U+311B) = oo（與 ㆦ 同音，鍵盤 'i' 鍵）
    _OO = 'ㄛ'
    _OO_FINALS = {
        _OO:                   'oo',
        _OO + 'ㄏ':        'ooh',   # ㄛㄏ → ooh
        _OO + 'ㄍ':        'ok',    # ㄛㄍ → ok
        _OO + 'ㄉ':        'oot',   # ㄛㄉ → oot（若有此音）
    }
    fin_map.update({k: v for k, v in _OO_FINALS.items() if k not in fin_map})

    # ㄇ (U+3107) 作為韻母 = 成節鼻音 m（對應 ㆬ）
    fin_map.setdefault('ㄇ',           'm')
    fin_map.setdefault('ㄇㄏ',     'mh')

    # ㄥ (U+3125) 單獨作為韻母 = ng
    fin_map.setdefault('ㄥ',           'ng')

    # ㄧㆤ + 入聲尾（iek / ieh）
    fin_map.setdefault('ㄧㆤㄍ', 'iek')   # ㄧㆤㄍ → iek
    fin_map.setdefault('ㄧㆤㄏ', 'ieh')   # ㄧㆤㄏ → ieh

    # ── 台灣延伸字元 → 標準注音別名 ──────────────────────────────────────────
    # 韻母表 key 用台灣延伸字元，漢字庫資料卻用標準注音字元，需加入標準形式別名
    # ㆬ (U+31AC, 台灣韻尾 m) ↔ ㄇ (U+3107, 標準 m)
    # ㆭ (U+31AD, 台灣韻尾 ng) ↔ ㄥ (U+3125, 標準 eng)
    # ㆦ (U+31A6, 台灣 oo) ↔ ㄛ (U+311B, 標準 o)
    EXT_TO_STD = {'ㆬ': 'ㄇ', 'ㆭ': 'ㄥ', 'ㆦ': 'ㄛ'}
    for tps_key, bpm2 in list(fin_map.items()):
        variant = ''.join(EXT_TO_STD.get(c, c) for c in tps_key)
        if variant != tps_key and variant not in fin_map:
            fin_map[variant] = bpm2

    # ── ㄧ + 入聲尾（直接使用標準字元，韻母表用特殊字元 一 U+4E00） ────────
    fin_map.setdefault('ㄧㄅ', 'ip')    # ㄧ + p-final
    fin_map.setdefault('ㄧㄉ', 'it')    # ㄧ + t-final
    fin_map.setdefault('ㄧㄍ', 'ik')    # ㄧ + k-final

    # ── ㄧㄨ + 入聲尾 ──────────────────────────────────────────────────────────
    fin_map.setdefault('ㄧㄨㄏ', 'iuh')    # iu + h-final

    # ── ㆪ (inn) + 入聲尾 ─────────────────────────────────────────────────────
    fin_map.setdefault('ㆪㄏ', 'innh')     # inn + h-final

    # ── ㄛ (oo) + 入聲尾 ──────────────────────────────────────────────────────
    fin_map.setdefault('ㄛㄅ', 'oop')      # oo + p-final

    # ── ㆯ (iaunn) 單獨出現 ───────────────────────────────────────────────────
    fin_map.setdefault('ㆯ', 'iaunn')

    # ── ㆨ (BOPOMOFO LETTER IR, U+31A8) → 對應 BPM2 「u」系列 ────────────────
    _IR = 'ㆨ'
    for _tps, _bpm2 in [
        (_IR,              'u'),
        (_IR + 'ㆤ',      'ue'),
        (_IR + 'ㆪ',      'uinn'),
        (_IR + 'ㄣ',      'un'),
        (_IR + 'ㆤㄏ',   'ueh'),
        (_IR + 'ㄣㄉ',   'und'),
        (_IR + 'ㄧ',     'ui'),      # ㆨㄧ → ui
        (_IR + 'ㄏ',     'uh'),      # ㆨㄏ → uh
    ]:
        fin_map.setdefault(_tps, _bpm2)

    # ── ㄧㄛ (io / ioo 系列) ──────────────────────────────────────────────────
    fin_map.setdefault('ㄧㄛ', 'io')           # io (短 o)
    fin_map.setdefault('ㄧㄛㄧ', 'ioi')        # ioi (如有)
    fin_map.setdefault('ㄧㆱ', 'ioom')         # i + oom

    # ── ㄧㄨㆱ / ㄛㆱ (ium / oom 系列) ──────────────────────────────────────
    fin_map.setdefault('ㄧㄨㆱ', 'ium')        # i+u+oom = ium
    fin_map.setdefault('ㄛㆱ',  'oom')         # oo + oom coda = oom

    # ── ㄨㄚㄍ / ㄨㆪㄏ / ㆮㄏ 入聲尾別名 ────────────────────────────────────
    fin_map.setdefault('ㄨㄚㄍ', 'uak')        # ua + k-final
    fin_map.setdefault('ㄨㆪㄏ', 'uinnh')      # uinn + h-final
    fin_map.setdefault('ㆮㄏ',   'ainnh')      # ainn + h-final

    # ── ㆯ / ㄧㆯ + 入聲尾 ────────────────────────────────────────────────────
    fin_map.setdefault('ㆯㄏ',  'iaunnh')      # iaunn + h-final
    fin_map.setdefault('ㄧㆯㄏ','iaunnh')      # i + iaunn + h-final

    # ── 空韻母（成節輔音：ㄇ/m 後無韻母）────────────────────────────────────
    fin_map.setdefault('', '')                  # 空韻母 → 允許成節輔音
    fin_map.setdefault('ㄏ', 'h')              # 單獨 h 入聲尾（如 mh）

    return siann_map, siann_set, fin_map


# ── 4. 單一音節轉換 ───────────────────────────────────────────────────────────
def convert_syllable(
    tps_str: str,
    siann_set: set,
    siann_map: dict,
    fin_map: dict,
) -> str | None:
    """
    將一個 TPS 音節字串轉換為 BPM2 字串。
    失敗時回傳 None。
    """
    s = _normalize_tps(tps_str.strip())
    if not s:
        return None

    # ── 取調號 ──
    tone = '1'
    last = s[-1]
    if last in TONE_SUFFIX:
        tone = TONE_SUFFIX[last]
        s = s[:-1]              # 去掉調號字元
    elif last in JIP_SIANN_UN_BUE:
        tone = '4'              # 入聲尾 → 保留字元（屬於韻母的一部份）

    if not s:
        return None

    # ── 拆分聲母 / 韻母 ──
    if s[0] in siann_set:
        siann_tps = s[0]
        un_tps    = s[1:]
    else:
        siann_tps = ''
        un_tps    = s

    # ── 查表 ──
    siann_bpm2 = siann_map.get(siann_tps, '') if siann_tps else ''
    un_bpm2    = fin_map.get(un_tps)

    if un_bpm2 is None:
        return None

    return siann_bpm2 + un_bpm2 + tone


# ── 5. 按鍵對映表 & 鍵碼 → TPS ───────────────────────────────────────────────
def build_key_map(path: str) -> dict[str, str]:
    """
    讀取【按鍵對映】工作表，建立 {鍵盤字串 → TPS字元} 對映。
    工作表格式：col B = "key_seq TPS_char"（如 ", ㆤ" 或 "ss ㄫ"）
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb['按鍵對映']
    key_map: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        # rsplit(' ', 1) 以最後一個空格分隔，支援 key 含多個字元
        parts = str(row[1]).strip().rsplit(' ', 1)
        if len(parts) == 2:
            key_str, tps_char = parts[0].strip(), parts[1].strip()
            if key_str and tps_char:
                key_map[key_str] = tps_char
    wb.close()
    return key_map


def keys_to_tps(key_seq: str, key_map: dict[str, str],
                sorted_keys: list) -> str | None:
    """
    將按鍵序列（如 '185'）貪婪匹配轉換為 TPS 字串（如 'ㄅㄚ˫'）。
    若有未知按鍵則回傳 None。
    """
    result = []
    i = 0
    while i < len(key_seq):
        matched = False
        for k in sorted_keys:          # 長度由長到短，優先匹配多字元鍵
            if key_seq[i:i + len(k)] == k:
                result.append(key_map[k])
                i += len(k)
                matched = True
                break
        if not matched:
            return None                 # 遇到未知按鍵
    return ''.join(result)


# ── 6. 主程式 ─────────────────────────────────────────────────────────────────
def main() -> None:
    import os
    default_path = os.path.join(os.path.dirname(__file__), 'CIN方音符號.xlsx')
    src_path = sys.argv[1] if len(sys.argv) > 1 else default_path
    dst_path = sys.argv[2] if len(sys.argv) > 2 else src_path

    print(f'讀取：{src_path}')
    siann_map, siann_set, fin_map = build_tables(src_path)
    print(f'  聲母對映：{len(siann_map)} 筆')
    print(f'  韻母對映：{len(fin_map)} 筆')

    # 按鍵對映：鍵盤字串 → TPS 字元（多字元鍵如 'ss'/'jj' 優先匹配）
    key_map = build_key_map(src_path)
    sorted_keys = sorted(key_map.keys(), key=len, reverse=True)
    print(f'  按鍵對映：{len(key_map)} 筆')

    # 載入工作簿（保留公式），只改寫 F 欄
    wb = load_workbook(src_path)
    ws = wb['漢字庫']

    converted = 0
    skipped   = 0   # A 欄空白或格式不符
    failed    = 0   # 無法對映
    fail_all  = []  # 所有失敗記錄 [(row_idx, key_seq, tps_or_None)]

    total_rows = ws.max_row - 3          # 資料列總數（第 4 列起）
    PROGRESS_STEP = 1000                 # 每 1000 列顯示一次進度

    print(f'\n開始轉換（共約 {total_rows:,} 列）…')
    for row_idx in range(4, ws.max_row + 1):
        # ── 進度顯示 ──
        processed = row_idx - 3
        if processed % PROGRESS_STEP == 0 or row_idx == ws.max_row:
            pct_done = processed / total_rows * 100
            print(f'  處理中… row {row_idx:>6,}  ({pct_done:5.1f}%)',
                  end='\r', flush=True)

        # ── 讀取 A 欄：格式為 "{key_seq} {漢字}" ──
        cell_a = ws.cell(row=row_idx, column=1).value
        if cell_a is None:
            skipped += 1
            continue
        raw_a = str(cell_a).strip()
        space_pos = raw_a.find(' ')
        if space_pos <= 0:
            skipped += 1
            continue
        key_seq = raw_a[:space_pos]

        # ── 鍵碼 → TPS ──
        tps_val = keys_to_tps(key_seq, key_map, sorted_keys)
        if tps_val is None:
            fail_all.append((row_idx, key_seq, None))
            failed += 1
            continue

        # ── TPS → BPM2 ──
        bpm2 = convert_syllable(tps_val, siann_set, siann_map, fin_map)
        if bpm2 is not None:
            ws.cell(row=row_idx, column=6).value = bpm2   # F 欄
            converted += 1
        else:
            fail_all.append((row_idx, key_seq, tps_val))
            failed += 1

    print()   # 換行，結束進度列
    wb.save(dst_path)

    total = converted + failed
    pct   = converted / total * 100 if total else 0
    print(f'\n轉換完成：')
    print(f'  成功 {converted:,} 筆 / 失敗 {failed:,} 筆（涵蓋率 {pct:.1f}%）')
    print(f'  空白略過 {skipped:,} 筆')
    if fail_all:
        print(f'\n失敗明細（共 {len(fail_all)} 筆）：')
        for ridx, ks, tps in fail_all:
            tps_s = repr(tps) if tps else '（鍵碼無法對映）'
            print(f'  row {ridx:>6,}  key={ks!r}  TPS={tps_s}')
    print(f'\n輸出：{dst_path}')


if __name__ == '__main__':
    main()
