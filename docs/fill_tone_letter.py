"""
fill_tone_letter.py
===================
根據響度優先次第：a > oo > (e = o) > (i = u) > 鼻音韻尾
在 Q 欄（Q4 至表格底端）寫入 Excel 公式，
讓試算表能動態計算「承載聲調符號的拼音字母」。

公式參照 C 欄（韻母原始資料），使用相對位址，向下自動調整。
"""

import openpyxl

WORKBOOK = r"C:\Users\AlanJui\work\rime-tlpa\docs\103_RIME對照工具【台語注音二式】.xlsx"
SHEET    = "【韻】韻母清單（台語注音二式==>閩拼方案）"
COL_F    = 6    # F 欄：閩拼方案韻母（根源資料，N=F、P=CONCAT(N,"(\d)")）
COL_Q    = 17   # Q 欄：標調字母（寫入公式）
START    = 4    # Row 4


def tone_formula(row: int) -> str:
    """
    產生第 row 列的 Excel 公式。
    響度優先次第：a > oo > (e = o，先出現者) > (i = u，先出現者) > 鼻音(m, ng)
    參照 F 欄（閩拼方案韻母根源資料）。
    """
    c = f"F{row}"
    return (
        f'=IF(ISNUMBER(FIND("a",{c})),"a",'
        f'IF(ISNUMBER(FIND("oo",{c})),"o",'
        f'IF(AND(ISNUMBER(FIND("e",{c})),ISNUMBER(FIND("o",{c}))),'
        f'IF(FIND("e",{c})<FIND("o",{c}),"e","o"),'
        f'IF(ISNUMBER(FIND("e",{c})),"e",'
        f'IF(ISNUMBER(FIND("o",{c})),"o",'
        f'IF(AND(ISNUMBER(FIND("i",{c})),ISNUMBER(FIND("u",{c}))),'
        f'IF(FIND("i",{c})<FIND("u",{c}),"i","u"),'
        f'IF(ISNUMBER(FIND("i",{c})),"i",'
        f'IF(ISNUMBER(FIND("u",{c})),"u",'
        f'IF(LEFT({c},1)="m","m",'
        f'IF(ISNUMBER(FIND("n",{c})),"n","")'
        f'))))))))))'
    )


def main():
    wb_read = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws_read = wb_read[SHEET]

    wb_edit = openpyxl.load_workbook(WORKBOOK)
    ws_edit = wb_edit[SHEET]

    # 找到 F 欄最後一筆韻母資料（純拉丁字母列）的列號
    last_row = START
    for r in range(START, ws_read.max_row + 1):
        val = ws_read.cell(r, COL_F).value
        if val and str(val).strip().isascii() and str(val).strip().isalpha():
            last_row = r

    print(f"寫入公式範圍：Q{START}:Q{last_row}\n")
    for r in range(START, last_row + 1):
        f_val = ws_read.cell(r, COL_F).value
        if not (f_val and str(f_val).strip().isascii() and str(f_val).strip().isalpha()):
            continue

        formula = tone_formula(r)
        old_q   = ws_read.cell(r, COL_Q).value
        ws_edit.cell(r, COL_Q).value = formula
        print(f"  Q{r} | F={str(f_val):<12s} | 原值={old_q!r:<5} | 公式={formula}")

    wb_edit.save(WORKBOOK)
    print(f"\n已儲存：{WORKBOOK}")


if __name__ == "__main__":
    main()
